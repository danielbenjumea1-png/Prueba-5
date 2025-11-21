import streamlit as st
import pandas as pd
import numpy as np
from pyzxing import BarCodeReader
from PIL import Image
import re
import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Colores para Excel
COLOR_VERDE = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
COLOR_MORADO = PatternFill(start_color="800080", end_color="800080", fill_type="solid")

st.title("📚 Inventario Biblioteca UCC - Sede Medellín")
st.write("La aplicación detecta códigos de barras automáticamente y actualiza el Excel sin necesidad de presionar botones.")

EXCEL_PATH = "inventario.xlsx"
BACKUP_PATH = "inventario_backup.xlsx"

# Función para crear backup
def crear_backup():
    if os.path.exists(EXCEL_PATH):
        shutil.copy(EXCEL_PATH, BACKUP_PATH)

if not os.path.exists(EXCEL_PATH):
    st.error("No se encontró 'inventario.xlsx'. Sube tu inventario inicial.")
    uploaded_file = st.file_uploader("Sube el inventario inicial", type=["xlsx"])
    if uploaded_file:
        with open(EXCEL_PATH, "wb") as f:
            f.write(uploaded_file.getbuffer())
        st.success("Inventario cargado exitosamente. Recarga la página para comenzar.")
    st.stop()

wb = load_workbook(EXCEL_PATH)
sheet = wb.active
df = pd.read_excel(EXCEL_PATH)

codigo_columna = None
for col in df.columns:
    if "codigo" in col.lower():
        codigo_columna = col
        break

if not codigo_columna:
    st.error("No existe una columna llamada 'codigo' en el archivo.")
    st.stop()

codigo_a_fila = {str(row[codigo_columna]).strip(): idx + 2 for idx, row in df.iterrows()}

# Inicializar lector de códigos de barras con pyzxing (pura Python, compatible con cloud)
reader = BarCodeReader()

st.subheader("Escanea el código de barras")
img_file = st.camera_input("Toma una foto del código de barras")

codigo_detectado = None

if img_file:
    # Cargar imagen con PIL
    img = Image.open(img_file)

    # Intentar decodificar códigos de barras con pyzxing
    results = reader.decode(img)

    posibles_codigos = []

    if results:
        for result in results:
            if 'parsed' in result and result['parsed']:
                data = result['parsed'].strip().upper()
                if data.startswith("B") and len(data) >= 7:
                    posibles_codigos.append(data)

    if posibles_codigos:
        codigo_detectado = max(posibles_codigos, key=len)
        st.success(f"Código de barras detectado: **{codigo_detectado}**")

        if codigo_detectado in codigo_a_fila:
            fila = codigo_a_fila[codigo_detectado]
            celda = f"A{fila}"
            sheet[celda].fill = COLOR_VERDE
            sheet[celda].font = Font(bold=True)
            st.success(f"✔ Código {codigo_detectado} encontrado y marcado en verde.")
        else:
            nueva_fila = sheet.max_row + 1
            sheet[f"A{nueva_fila}"] = codigo_detectado
            sheet[f"A{nueva_fila}"].fill = COLOR_MORADO
            sheet[f"A{nueva_fila}"].font = Font(bold=True)
            # Actualizar el mapeo para futuros escaneos
            codigo_a_fila[codigo_detectado] = nueva_fila
            st.warning(f"➕ Código nuevo agregado: {codigo_detectado}")

        wb.save(EXCEL_PATH)
        # Crear backup
        crear_backup()

    else:
        st.warning("No se encontró un código de barras válido en la imagen. Intenta con OCR si es texto impreso.")
        
st.subheader("Ingresar código manualmente")

if 'codigo_manual' not in st.session_state:
    st.session_state['codigo_manual'] = ''

codigo_manual = st.text_input("Escribe el código si no puedes escanearlo:", value=st.session_state['codigo_manual'])

if st.button("Procesar Código Manual"):
    if codigo_manual:
        codigo_manual = codigo_manual.strip().upper()

        if codigo_manual in codigo_a_fila:
            fila = codigo_a_fila[codigo_manual]
            celda = f"A{fila}"
            sheet[celda].fill = COLOR_VERDE
            sheet[celda].font = Font(bold=True)
            st.success(f"✔ Código {codigo_manual} encontrado y marcado en verde.")

        else:
            nueva_fila = sheet.max_row + 1
            sheet[f"A{nueva_fila}"] = codigo_manual
            sheet[f"A{nueva_fila}"].fill = COLOR_MORADO
            sheet[f"A{nueva_fila}"].font = Font(bold=True)
            codigo_a_fila[codigo_manual] = nueva_fila
            st.warning(f"➕ Código nuevo agregado manualmente: {codigo_manual}")

        wb.save(EXCEL_PATH)
        crear_backup()
        st.session_state['codigo_manual'] = ''  # Resetear el input
        st.rerun()  # Forzar recarga para vaciar el campo inmediatamente
    else:
        st.warning("Por favor, ingresa un código antes de procesar.")
    
st.subheader("Inventario actualizado")
st.dataframe(pd.read_excel(EXCEL_PATH))

with open(EXCEL_PATH, "rb") as f:
    st.download_button("Descargar inventario actualizado", f, file_name="inventario_actualizado.xlsx")
